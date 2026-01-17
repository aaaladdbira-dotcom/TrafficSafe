"""
Data Export Service
===================
Export data in various formats (CSV, Excel, PDF)
"""

import io
import csv
from datetime import datetime
from flask import Response, send_file


def export_to_csv(data, filename, headers=None):
    """
    Export data to CSV format.
    
    Args:
        data: List of dicts or list of lists
        filename: Output filename
        headers: Optional list of headers (required if data is list of lists)
    
    Returns:
        Flask Response with CSV file
    """
    # Build CSV text in memory and return as an attachment using send_file
    output = io.StringIO()

    if data and isinstance(data[0], dict):
        # Dict data - use keys as headers
        if not headers:
            headers = list(data[0].keys())
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        writer.writerows(data)
    else:
        # List data
        writer = csv.writer(output)
        if headers:
            writer.writerow(headers)
        writer.writerows(data)

    # Convert text to bytes for send_file
    output_bytes = io.BytesIO()
    output_bytes.write(output.getvalue().encode('utf-8'))
    output_bytes.seek(0)
    output.close()

    return send_file(
        output_bytes,
        as_attachment=True,
        download_name=filename,
        mimetype='text/csv'
    )


def export_to_excel(data, filename, headers=None, sheet_name='Data'):
    """
    Export data to Excel format.
    
    Args:
        data: List of dicts or list of lists
        filename: Output filename
        headers: Optional list of headers
        sheet_name: Excel sheet name
    
    Returns:
        Flask Response with Excel file
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        # Fallback to CSV if openpyxl not available
        return export_to_csv(data, filename.replace('.xlsx', '.csv'), headers)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get headers
    if data and isinstance(data[0], dict):
        if not headers:
            headers = list(data[0].keys())
        rows = [[row.get(h, '') for h in headers] for row in data]
    else:
        rows = data
    
    # Write headers
    if headers:
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
    
    # Write data
    start_row = 2 if headers else 1
    for row_idx, row in enumerate(rows, start_row):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
    
    # Save to bytes and return as attachment
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def export_to_pdf(data, filename, title='Report', headers=None):
    """
    Export data to PDF format.
    
    Args:
        data: List of dicts or list of lists
        filename: Output filename
        title: PDF title
        headers: Optional list of headers
    
    Returns:
        Flask Response with PDF file
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        # Fallback to CSV if reportlab not available
        return export_to_csv(data, filename.replace('.pdf', '.csv'), headers)
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(letter),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20
    )
    elements.append(Paragraph(title, title_style))
    
    # Subtitle with date
    subtitle = f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    elements.append(Paragraph(subtitle, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Get headers and rows
    if data and isinstance(data[0], dict):
        if not headers:
            headers = list(data[0].keys())
        rows = [[row.get(h, '') for h in headers] for row in data]
    else:
        rows = data
    
    # Create table data
    table_data = []
    if headers:
        table_data.append(headers)
    table_data.extend(rows)
    
    # Create table
    table = Table(table_data, repeatRows=1)
    
    # Table style
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ])
    
    table.setStyle(style)
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )


def format_accident_for_export(accident):
    """Format an accident record for export"""
    return {
        'ID': accident.id,
        'Date': accident.occurred_at.strftime('%Y-%m-%d %H:%M') if accident.occurred_at else '',
        'Location': accident.location or '',
        'Governorate': accident.governorate or '',
        'Delegation': accident.delegation or '',
        'Severity': accident.severity or '',
        'Cause': accident.cause or '',
        'Source': accident.source or '',
        'Created At': accident.created_at.strftime('%Y-%m-%d %H:%M') if accident.created_at else ''
    }


def format_user_for_export(user):
    """Format a user record for export"""
    return {
        'ID': user.id,
        'Name': user.full_name or '',
        'Email': user.email or '',
        'Role': user.role or '',
        'User Type': user.user_type or '',
        'Gender': user.gender or '',
        'Work Place': user.work_place or ''
    }
